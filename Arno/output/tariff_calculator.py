#!/usr/bin/env python3
"""
tariff_calculator.py

Main program to read inputs from the Excel tariff calculator and print:
- Premium Calculation outputs
- All Progression Values rows

Reads from:
  file: Tariff_Calculator.xlsm
  sheet: Calculation

Cell mapping is based on the provided screenshot.
"""

from __future__ import annotations

from dataclasses import asdict
from typing import Any, Dict, List

from openpyxl import load_workbook

from premium_and_progress_values import (
    PolicyInputs,
    TariffInputs,
    Limits,
    calc_premium_calculation,
    calc_progression_values,
)


EXCEL_FILE = "Tariff_Calculator.xlsm"
SHEET_NAME = "Calculation"


def _read(ws, addr: str) -> Any:
    v = ws[addr].value
    if v is None:
        raise ValueError(f"Cell {addr} is empty (sheet '{ws.title}').")
    return v


def _as_float(v: Any) -> float:
    # openpyxl returns numbers as int/float already; Excel percentages are typically floats (e.g., 0.0175).
    try:
        return float(v)
    except Exception as e:
        raise ValueError(f"Expected a number, got {v!r}") from e


def _as_int(v: Any) -> int:
    try:
        return int(v)
    except Exception as e:
        raise ValueError(f"Expected an integer, got {v!r}") from e


def _as_str(v: Any) -> str:
    return str(v).strip()


def _print_kv(title: str, d: Dict[str, Any]) -> None:
    print(f"\n{title}")
    print("-" * len(title))
    for k, v in d.items():
        if isinstance(v, float):
            print(f"{k:>20s}: {v:,.8f}")
        else:
            print(f"{k:>20s}: {v}")


def _print_table(rows: List[Dict[str, Any]], cols: List[str]) -> None:
    # simple fixed-width table printer
    widths = {c: max(len(c), max(len(f"{r.get(c, '')}") for r in rows)) for c in cols}
    header = " | ".join(c.ljust(widths[c]) for c in cols)
    sep = "-+-".join("-" * widths[c] for c in cols)
    print("\nProgression values")
    print("------------------")
    print(header)
    print(sep)
    for r in rows:
        line = " | ".join(f"{r.get(c, '')}".ljust(widths[c]) for c in cols)
        print(line)


def main() -> None:
    wb = load_workbook(EXCEL_FILE, data_only=True, keep_vba=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Worksheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    # --- Inputs (from screenshot) ---
    # Policy data (A4:A9 labels, B4:B9 values)
    policy = PolicyInputs(
        x=_as_int(_read(ws, "B4")),
        sex=_as_str(_read(ws, "B5")),
        n=_as_int(_read(ws, "B6")),
        t=_as_int(_read(ws, "B7")),
        sum_insured=_as_float(_read(ws, "B8")),
        pay_freq=_as_int(_read(ws, "B9")),
    )

    # Tariff data (D4:D12 labels, E4:E12 values)
    tariff = TariffInputs(
        interest_rate=_as_float(_read(ws, "E4")),
        mortality_table=_as_str(_read(ws, "E5")),
        alpha=_as_float(_read(ws, "E6")),
        beta1=_as_float(_read(ws, "E7")),
        gamma1=_as_float(_read(ws, "E8")),
        gamma2=_as_float(_read(ws, "E9")),
        gamma3=_as_float(_read(ws, "E10")),
        k=_as_float(_read(ws, "E11")),
        modal_surcharge=_as_float(_read(ws, "E12")),
    )

    # Limits (G4:G5 labels, H4:H5 values)
    limits = Limits(
        min_age_flex=_as_int(_read(ws, "H4")),
        min_term_flex=_as_int(_read(ws, "H5")),
    )

    # --- Calculations ---
    premium = calc_premium_calculation(policy, tariff)
    progression = calc_progression_values(
        policy,
        tariff,
        limits,
        pxt=premium["Pxt"],
        gross_annual_prem=premium["GrossAnnualPrem"],
        max_k=policy.n,  # "all progress values" -> k = 0..n
    )

    # --- Output ---
    _print_kv("Inputs: Policy", asdict(policy))
    _print_kv("Inputs: Tariff", asdict(tariff))
    _print_kv("Inputs: Limits", asdict(limits))

    _print_kv("Premium calculation", premium)

    cols = [
        "k",
        "Axn",
        "axn",
        "axt",
        "kVx_pp",
        "kDRx_pp",
        "kVx_pu",
        "kVx_MRV",
        "Flex_phase",
        "Surrender_deduction",
        "Surrender_value",
        "SumInsured_pu",
    ]
    # format floats a bit for readability
    pretty_rows: List[Dict[str, Any]] = []
    for r in progression:
        pr = dict(r)
        for c in cols:
            if c != "k" and isinstance(pr.get(c), float):
                pr[c] = f"{pr[c]:.8f}"
        pretty_rows.append(pr)

    _print_table(pretty_rows, cols)


if __name__ == "__main__":
    main()
