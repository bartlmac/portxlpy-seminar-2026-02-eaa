#!/usr/bin/env python3
"""
compare.py

Compares Python-calculated tariff results against the values in the Excel model.

- Reads inputs from: Tariff_Calculator.xlsm, sheet "Calculation"
- Computes:
    * Premium Calculation block
    * Progression Values table (k = 0..n)
- Reads corresponding Excel outputs and checks they match within tolerances.
- Prints a clear mismatch report and a final PASS/FAIL summary.

Cell mapping is based on the provided screenshot.

Assumptions:
- premium_and_progress_values.py exists and is importable.
- presentvalues.py / commvalues.py exist and implement the actuarial functions.
"""

from __future__ import annotations

from dataclasses import asdict
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from premium_and_progress_values import (
    PolicyInputs,
    TariffInputs,
    Limits,
    calc_premium_calculation,
    calc_progression_values,
)

import sys

EXCEL_FILE = "Tariff_Calculator.xlsm"
SHEET_NAME = "Calculation"

# Comparison tolerances
ABS_TOL = 1e-8
REL_TOL = 1e-8


def _read(ws, addr: str) -> Any:
    return ws[addr].value


def _req(ws, addr: str) -> Any:
    v = _read(ws, addr)
    if v is None:
        raise ValueError(f"Cell {addr} is empty (sheet '{ws.title}').")
    return v


def _as_float(v: Any) -> float:
    try:
        return float(v)
    except Exception as e:
        raise ValueError(f"Expected numeric, got {v!r}") from e


def _as_int(v: Any) -> int:
    try:
        return int(v)
    except Exception as e:
        raise ValueError(f"Expected int, got {v!r}") from e


def _as_str(v: Any) -> str:
    return str(v).strip()


def approx_equal(a: float, b: float, abs_tol: float = ABS_TOL, rel_tol: float = REL_TOL) -> bool:
    diff = abs(a - b)
    if diff <= abs_tol:
        return True
    scale = max(abs(a), abs(b), 1.0)
    return diff <= rel_tol * scale


def compare_value(name: str, py: float, xl: Any) -> Tuple[bool, str]:
    if xl is None:
        return False, f"{name}: Excel value is empty"
    try:
        xl_f = float(xl)
    except Exception:
        return False, f"{name}: Excel value is non-numeric ({xl!r})"
    ok = approx_equal(py, xl_f)
    if ok:
        return True, ""
    diff = py - xl_f
    return False, f"{name}: PY={py:.12g}  XL={xl_f:.12g}  DIFF={diff:.12g}"


def main() -> None:
    wb = load_workbook(EXCEL_FILE, data_only=True, keep_vba=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Worksheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    # ---- Inputs (from screenshot) ----
    policy = PolicyInputs(
        x=_as_int(_req(ws, "B4")),
        sex=_as_str(_req(ws, "B5")),
        n=_as_int(_req(ws, "B6")),
        t=_as_int(_req(ws, "B7")),
        sum_insured=_as_float(_req(ws, "B8")),
        pay_freq=_as_int(_req(ws, "B9")),
    )

    tariff = TariffInputs(
        interest_rate=_as_float(_req(ws, "E4")),
        mortality_table=_as_str(_req(ws, "E5")),
        alpha=_as_float(_req(ws, "E6")),
        beta1=_as_float(_req(ws, "E7")),
        gamma1=_as_float(_req(ws, "E8")),
        gamma2=_as_float(_req(ws, "E9")),
        gamma3=_as_float(_req(ws, "E10")),
        k=_as_float(_req(ws, "E11")),
        modal_surcharge=_as_float(_req(ws, "E12")),
    )

    limits = Limits(
        min_age_flex=_as_int(_req(ws, "H4")),
        min_term_flex=_as_int(_req(ws, "H5")),
    )

    # ---- Python calculation ----
    prem_py = calc_premium_calculation(policy, tariff)
    prog_py = calc_progression_values(
        policy,
        tariff,
        limits,
        pxt=prem_py["Pxt"],
        gross_annual_prem=prem_py["GrossAnnualPrem"],
        max_k=policy.n,
    )

    # ---- Excel outputs mapping (corrected) ----
    prem_xl_cells = {
        "NormGrossAnnualPrem": "K5",
        "GrossAnnualPrem": "K6",
        "GrossModalPrem": "K7",
        "Pxt": "K9",   # <-- corrected cell reference
    }

    # Progression table layout
    prog_cols = [
        ("k", "A"),
        ("Axn", "B"),
        ("axn", "C"),
        ("axt", "D"),
        ("kVx_pp", "E"),
        ("kDRx_pp", "F"),
        ("kVx_pu", "G"),
        ("kVx_MRV", "H"),
        ("Flex_phase", "I"),
        ("Surrender_deduction", "J"),
        ("Surrender_value", "K"),
        ("SumInsured_pu", "L"),
    ]
    start_row = 16

    mismatches: List[str] = []

    # Premium block comparisons
    for name, addr in prem_xl_cells.items():
        ok, msg = compare_value(name, float(prem_py[name]), _read(ws, addr))
        if not ok:
            mismatches.append(msg)

    # Progression table comparisons
    for idx, r_py in enumerate(prog_py):
        row = start_row + idx
        k_xl = _read(ws, f"A{row}")
        if k_xl is None:
            mismatches.append(f"Progression row {row}: Excel k is empty (expected {r_py['k']})")
            continue
        try:
            if int(k_xl) != int(r_py["k"]):
                mismatches.append(f"Progression row {row}: k mismatch PY={r_py['k']} XL={k_xl}")
        except Exception:
            mismatches.append(f"Progression row {row}: k non-integer in Excel ({k_xl!r})")

        for name, col in prog_cols:
            addr = f"{col}{row}"
            xl_val = _read(ws, addr)

            if name in ("k", "Flex_phase"):
                if xl_val is None:
                    mismatches.append(f"{name} @ {addr}: Excel empty")
                    continue
                try:
                    py_i = int(r_py[name])
                    xl_i = int(xl_val)
                    if py_i != xl_i:
                        mismatches.append(f"{name} @ {addr}: PY={py_i} XL={xl_i}")
                except Exception:
                    mismatches.append(f"{name} @ {addr}: non-integer compare (PY={r_py[name]!r}, XL={xl_val!r})")
                continue

            ok, msg = compare_value(f"{name} @ {addr}", float(r_py[name]), xl_val)
            if not ok:
                mismatches.append(msg)

    # ---- Reporting ----
    print("Inputs used (from Excel):")
    print("  Policy:", asdict(policy))
    print("  Tariff:", asdict(tariff))
    print("  Limits:", asdict(limits))

    print("\nPremium calculation (Python):")
    for k, v in prem_py.items():
        print(f"  {k:>20s}: {v:.12g}")

    return 1 if mismatches else 0


if __name__ == "__main__":
    sys.exit(main())